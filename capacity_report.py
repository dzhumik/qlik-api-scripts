# -*- coding: utf-8 -*-
"""
Capacity License Report - Excel Output
Halyk Bank, DTPD - Self-Service BI

Создаёт Excel-файл с двумя листами:
1. "Присвоения" - история присвоений Analyzer-лицензий (помесячно)
2. "Потребление" - реальное потребление Capacity-минут (помесячно по пользователям)

Требования:
    pip install requests openpyxl
"""

import requests
from datetime import datetime
from pathlib import Path
import urllib3
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =====================================================================
# НАСТРОЙКИ
# =====================================================================

QLIK_SERVER = "fdata1a003.halykbank.nb"
QLIK_USER   = "UserDirectory=UNIVERSAL;UserId=00060961"

BASE_DIR    = Path(__file__).parent
CLIENT_CERT = str(BASE_DIR / "certificate" / "client.pem")
CLIENT_KEY  = str(BASE_DIR / "certificate" / "client_key.pem")
ROOT_CERT   = str(BASE_DIR / "certificate" / "root.pem")

XRFKEY = "0123456789abcdef"

# =====================================================================

HEADERS = {
    "X-Qlik-Xrfkey": XRFKEY,
    "X-Qlik-User": QLIK_USER,
    "Content-Type": "application/json"
}
PARAMS = {"xrfkey": XRFKEY}
QLIK_URL = f"https://{QLIK_SERVER}:4242/qrs"

MONTH_NAMES = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}

# Стили
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E78")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E78")
SECTION_FILL = PatternFill("solid", start_color="DDEBF7")
TOTAL_FONT = Font(name="Arial", size=11, bold=True)
TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


# =====================================================================
# QRS API
# =====================================================================

def parse_qlik_time(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except (ValueError, AttributeError):
        return None


def qrs_get(endpoint):
    url = f"{QLIK_URL}/{endpoint}"
    try:
        return requests.get(
            url, headers=HEADERS, params=PARAMS,
            cert=(CLIENT_CERT, CLIENT_KEY),
            verify=ROOT_CERT, timeout=120
        )
    except Exception as e:
        print(f"\n[ERROR] {e}")
        sys.exit(1)


def fetch_assignments():
    print("[1/4] Получаем историю присвоений Analyzer-лицензий...")
    response = qrs_get("license/analyzeraccesstype/full")
    if response.status_code != 200:
        print(f"  Ошибка {response.status_code}: {response.text[:300]}")
        sys.exit(1)
    data = response.json()
    print(f"  Записей о присвоениях: {len(data)}")
    return data


def fetch_usage():
    print("[2/4] Получаем данные о реальном потреблении Capacity...")
    response = qrs_get("license/analyzertimeaccessusage/full")
    if response.status_code != 200:
        print(f"  Ошибка {response.status_code}: {response.text[:300]}")
        sys.exit(1)
    data = response.json()
    print(f"  Записей о сессиях: {len(data)}")
    return data


# =====================================================================
# ПОДГОТОВКА ДАННЫХ
# =====================================================================

def prepare_assignments(records):
    """Возвращает список словарей: [{month, year, login, name, date}, ...]"""
    print("[3/4] Готовим данные присвоений...")
    rows = []
    for rec in records:
        user_obj = rec.get("user", {})
        if not isinstance(user_obj, dict):
            continue
        user_dir = user_obj.get("userDirectory", "")
        user_id = user_obj.get("userId", "")
        user_name = user_obj.get("name") or user_id
        login = f"{user_dir}\\{user_id}" if user_dir else user_id
        if not login or login == "\\":
            continue
        created = parse_qlik_time(rec.get("createdDate"))
        if not created:
            continue
        rows.append({
            "year": created.year,
            "month": created.month,
            "month_name": MONTH_NAMES[created.month],
            "login": login,
            "name": user_name,
            "date": created
        })
    rows.sort(key=lambda r: r["date"])
    return rows


def prepare_usage(records):
    """Возвращает агрегированный словарь: {(year, month): {login: {...}}}"""
    print("[4/4] Готовим данные потребления...")
    monthly = {}
    for rec in records:
        user_obj = rec.get("user", {})
        if not isinstance(user_obj, dict):
            continue
        user_dir = user_obj.get("userDirectory", "")
        user_id = user_obj.get("userId", "")
        user_name = user_obj.get("name") or user_id
        login = f"{user_dir}\\{user_id}" if user_dir else user_id
        if not login or login == "\\":
            continue
        start = parse_qlik_time(rec.get("useStartTime"))
        stop = parse_qlik_time(rec.get("useStopTime"))
        if not start:
            continue
        if stop and stop > start:
            minutes = (stop - start).total_seconds() / 60.0
        else:
            minutes = 6.0
        key = (start.year, start.month)
        if key not in monthly:
            monthly[key] = {}
        if login not in monthly[key]:
            monthly[key][login] = {
                "name": user_name,
                "minutes": 0.0,
                "sessions": 0,
                "first_login": None,
                "last_login": None
            }
        monthly[key][login]["minutes"] += minutes
        monthly[key][login]["sessions"] += 1
        if not monthly[key][login]["first_login"] or start < monthly[key][login]["first_login"]:
            monthly[key][login]["first_login"] = start
        if not monthly[key][login]["last_login"] or start > monthly[key][login]["last_login"]:
            monthly[key][login]["last_login"] = start
    return monthly


# =====================================================================
# СОЗДАНИЕ EXCEL
# =====================================================================

def style_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, alignment=LEFT):
    cell.font = DATA_FONT
    cell.alignment = alignment
    cell.border = THIN_BORDER


def style_total_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER


def build_assignments_sheet(ws, assignments):
    """Лист 1: История присвоений."""
    
    # Заголовок
    ws["A1"] = "История присвоений Analyzer-лицензий"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    
    ws["A2"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Всего записей: {len(assignments)}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A2:E2")
    
    # ===== Сводная таблица по месяцам =====
    ws["A4"] = "Сводка по месяцам"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:E4")
    
    headers_summary = ["Месяц", "Год", "Новых присвоений", "Накопительно", ""]
    row_idx = 5
    for col, h in enumerate(headers_summary, 1):
        ws.cell(row=row_idx, column=col, value=h)
    style_header_row(ws, row_idx, 4)
    
    # Группируем по (year, month)
    by_month = {}
    for a in assignments:
        key = (a["year"], a["month"])
        by_month.setdefault(key, 0)
        by_month[key] += 1
    
    cumulative = 0
    row_idx = 6
    for key in sorted(by_month.keys()):
        year, month = key
        count = by_month[key]
        cumulative += count
        
        ws.cell(row=row_idx, column=1, value=MONTH_NAMES[month])
        ws.cell(row=row_idx, column=2, value=year)
        ws.cell(row=row_idx, column=3, value=count)
        ws.cell(row=row_idx, column=4, value=cumulative)
        
        style_data_cell(ws.cell(row=row_idx, column=1), LEFT)
        style_data_cell(ws.cell(row=row_idx, column=2), CENTER)
        style_data_cell(ws.cell(row=row_idx, column=3), RIGHT)
        style_data_cell(ws.cell(row=row_idx, column=4), RIGHT)
        
        row_idx += 1
    
    # Итого
    ws.cell(row=row_idx, column=1, value="ВСЕГО")
    ws.cell(row=row_idx, column=3, value=cumulative)
    style_total_row(ws, row_idx, 4)
    
    # ===== Детальная таблица =====
    detail_start = row_idx + 3
    
    ws.cell(row=detail_start, column=1, value="Детальный список присвоений")
    ws.cell(row=detail_start, column=1).font = SECTION_FONT
    ws.cell(row=detail_start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=5)
    
    headers_detail = ["Месяц", "Год", "Логин", "ФИО", "Дата присвоения"]
    detail_header_row = detail_start + 1
    for col, h in enumerate(headers_detail, 1):
        ws.cell(row=detail_header_row, column=col, value=h)
    style_header_row(ws, detail_header_row, 5)
    
    # Включаем фильтры
    last_data_row = detail_header_row + len(assignments)
    ws.auto_filter.ref = f"A{detail_header_row}:E{last_data_row}"
    
    # Замораживаем
    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=1)
    
    row_idx = detail_header_row + 1
    for a in assignments:
        ws.cell(row=row_idx, column=1, value=a["month_name"])
        ws.cell(row=row_idx, column=2, value=a["year"])
        ws.cell(row=row_idx, column=3, value=a["login"])
        ws.cell(row=row_idx, column=4, value=a["name"])
        ws.cell(row=row_idx, column=5, value=a["date"].strftime("%d.%m.%Y %H:%M"))
        
        style_data_cell(ws.cell(row=row_idx, column=1), LEFT)
        style_data_cell(ws.cell(row=row_idx, column=2), CENTER)
        style_data_cell(ws.cell(row=row_idx, column=3), LEFT)
        style_data_cell(ws.cell(row=row_idx, column=4), LEFT)
        style_data_cell(ws.cell(row=row_idx, column=5), CENTER)
        
        row_idx += 1
    
    # Ширина колонок
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20


def build_usage_sheet(ws, usage):
    """Лист 2: Реальное потребление."""
    
    # Заголовок
    ws["A1"] = "Реальное потребление Analyzer Capacity"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    
    total_records = sum(len(users) for users in usage.values())
    ws["A2"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Уникальных пользователей-месяцев: {total_records}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A2:G2")
    
    # ===== Сводка по месяцам =====
    ws["A4"] = "Сводка по месяцам"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:G4")
    
    headers_summary = ["Месяц", "Год", "Активных пользователей", "Минут", "Часов", "Сессий", ""]
    row_idx = 5
    for col, h in enumerate(headers_summary, 1):
        ws.cell(row=row_idx, column=col, value=h)
    style_header_row(ws, row_idx, 6)
    
    grand_min = 0
    grand_ses = 0
    
    row_idx = 6
    for key in sorted(usage.keys()):
        year, month = key
        users = usage[key]
        total_min = sum(u["minutes"] for u in users.values())
        total_ses = sum(u["sessions"] for u in users.values())
        grand_min += total_min
        grand_ses += total_ses
        
        ws.cell(row=row_idx, column=1, value=MONTH_NAMES[month])
        ws.cell(row=row_idx, column=2, value=year)
        ws.cell(row=row_idx, column=3, value=len(users))
        ws.cell(row=row_idx, column=4, value=round(total_min, 1))
        ws.cell(row=row_idx, column=5, value=round(total_min / 60, 2))
        ws.cell(row=row_idx, column=6, value=total_ses)
        
        style_data_cell(ws.cell(row=row_idx, column=1), LEFT)
        style_data_cell(ws.cell(row=row_idx, column=2), CENTER)
        style_data_cell(ws.cell(row=row_idx, column=3), RIGHT)
        style_data_cell(ws.cell(row=row_idx, column=4), RIGHT)
        style_data_cell(ws.cell(row=row_idx, column=5), RIGHT)
        style_data_cell(ws.cell(row=row_idx, column=6), RIGHT)
        
        row_idx += 1
    
    # Итого
    ws.cell(row=row_idx, column=1, value="ВСЕГО")
    ws.cell(row=row_idx, column=4, value=round(grand_min, 1))
    ws.cell(row=row_idx, column=5, value=round(grand_min / 60, 2))
    ws.cell(row=row_idx, column=6, value=grand_ses)
    style_total_row(ws, row_idx, 6)
    
    # ===== Детальная таблица =====
    detail_start = row_idx + 3
    
    ws.cell(row=detail_start, column=1, value="Детальное потребление по пользователям")
    ws.cell(row=detail_start, column=1).font = SECTION_FONT
    ws.cell(row=detail_start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=9)
    
    headers_detail = ["Месяц", "Год", "Логин", "ФИО", "Минут", "Часов", "Сессий", "Первый вход", "Последний вход"]
    detail_header_row = detail_start + 1
    for col, h in enumerate(headers_detail, 1):
        ws.cell(row=detail_header_row, column=col, value=h)
    style_header_row(ws, detail_header_row, 9)
    
    # Считаем общее число строк для AutoFilter
    total_rows = sum(len(users) for users in usage.values())
    last_data_row = detail_header_row + total_rows
    ws.auto_filter.ref = f"A{detail_header_row}:I{last_data_row}"
    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=1)
    
    row_idx = detail_header_row + 1
    for key in sorted(usage.keys()):
        year, month = key
        sorted_users = sorted(usage[key].items(), key=lambda x: -x[1]["minutes"])
        for login, data in sorted_users:
            first = data["first_login"].strftime("%d.%m.%Y %H:%M") if data["first_login"] else ""
            last = data["last_login"].strftime("%d.%m.%Y %H:%M") if data["last_login"] else ""
            
            ws.cell(row=row_idx, column=1, value=MONTH_NAMES[month])
            ws.cell(row=row_idx, column=2, value=year)
            ws.cell(row=row_idx, column=3, value=login)
            ws.cell(row=row_idx, column=4, value=data["name"])
            ws.cell(row=row_idx, column=5, value=round(data["minutes"], 1))
            ws.cell(row=row_idx, column=6, value=round(data["minutes"] / 60, 2))
            ws.cell(row=row_idx, column=7, value=data["sessions"])
            ws.cell(row=row_idx, column=8, value=first)
            ws.cell(row=row_idx, column=9, value=last)
            
            style_data_cell(ws.cell(row=row_idx, column=1), LEFT)
            style_data_cell(ws.cell(row=row_idx, column=2), CENTER)
            style_data_cell(ws.cell(row=row_idx, column=3), LEFT)
            style_data_cell(ws.cell(row=row_idx, column=4), LEFT)
            style_data_cell(ws.cell(row=row_idx, column=5), RIGHT)
            style_data_cell(ws.cell(row=row_idx, column=6), RIGHT)
            style_data_cell(ws.cell(row=row_idx, column=7), RIGHT)
            style_data_cell(ws.cell(row=row_idx, column=8), CENTER)
            style_data_cell(ws.cell(row=row_idx, column=9), CENTER)
            
            row_idx += 1
    
    # Ширина колонок
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20


def save_excel(assignments, usage):
    wb = Workbook()
    
    # Лист 1: Присвоения
    ws1 = wb.active
    ws1.title = "Присвоения"
    build_assignments_sheet(ws1, assignments)
    
    # Лист 2: Потребление
    ws2 = wb.create_sheet("Потребление")
    build_usage_sheet(ws2, usage)
    
    filename = f"capacity_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = BASE_DIR / filename
    wb.save(filepath)
    
    print(f"\nExcel сохранён: {filepath}")
    return filepath


# =====================================================================
# MAIN
# =====================================================================

def main():
    print("=" * 80)
    print(f"  Capacity License Report (Excel)  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print(f"  Сервер: {QLIK_SERVER}")
    print("=" * 80)
    print()
    
    # Получаем данные
    assignment_records = fetch_assignments()
    usage_records = fetch_usage()
    
    # Готовим
    assignments = prepare_assignments(assignment_records)
    usage = prepare_usage(usage_records)
    
    # Сводка в консоль
    print()
    print("=" * 80)
    print("  СВОДКА")
    print("=" * 80)
    
    print(f"\nЛист 1 (Присвоения): {len(assignments)} записей")
    by_month_a = {}
    for a in assignments:
        key = (a["year"], a["month"])
        by_month_a.setdefault(key, 0)
        by_month_a[key] += 1
    print(f"  Уникальных месяцев: {len(by_month_a)}")
    print(f"  Период: {min(a['date'] for a in assignments).strftime('%m.%Y')} - {max(a['date'] for a in assignments).strftime('%m.%Y')}")
    
    print(f"\nЛист 2 (Потребление):")
    grand_min = 0
    grand_ses = 0
    total_users = set()
    for users in usage.values():
        grand_min += sum(u["minutes"] for u in users.values())
        grand_ses += sum(u["sessions"] for u in users.values())
        total_users.update(users.keys())
    print(f"  Уникальных активных пользователей: {len(total_users)}")
    print(f"  Всего минут: {grand_min:.1f} ({grand_min/60:.1f} часов)")
    print(f"  Всего сессий: {grand_ses}")
    
    print()
    save_excel(assignments, usage)
    
    print("\nГотово.")


if __name__ == "__main__":
    main()
